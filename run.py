import os, base64, gc, io, json, re, time, logging
import concurrent.futures, threading
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
import pdfplumber, anthropic, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

log = logging.getLogger(__name__)

# Cache: msg_id -> list of extracted PDF text strings
_pdf_text_cache = {}

# Thread-safe Gmail search cache and lock (Gmail API client is NOT thread-safe)
_gmail_lock = threading.Lock()
_gmail_search_cache = {}  # (id(svc), query) -> list of message dicts

MAX_RETRIES = 3
RETRY_DELAYS = [2, 5, 10]
MAX_WORKERS = 10  # parallel threads — tuned for 16GB RAM iMac
BATCH_SIZE = 10   # process invoices in batches to limit peak memory
GMAIL_MAX_RESULTS = 30  # max emails per search — must be high enough for frequent senders like Tasfresh


def _retry(fn, description="operation"):
    """Retry fn() on transient errors (rate limits, server errors, auth errors)."""
    for attempt in range(MAX_RETRIES):
        try:
            return fn()
        except anthropic.RateLimitError as e:
            delay = RETRY_DELAYS[attempt]
            log.warning(f"{description}: rate limited, retrying in {delay}s (attempt {attempt+1})")
            time.sleep(delay)
        except anthropic.InternalServerError as e:
            delay = RETRY_DELAYS[attempt]
            log.warning(f"{description}: server error, retrying in {delay}s (attempt {attempt+1})")
            time.sleep(delay)
        except HttpError as e:
            if e.resp.status in (401, 403):
                # Auth error — token may have expired mid-run, retry once
                log.warning(f"{description}: Gmail auth error {e.resp.status}, retrying (attempt {attempt+1})")
                time.sleep(1)
            elif e.resp.status in (429, 500, 503):
                delay = RETRY_DELAYS[attempt]
                log.warning(f"{description}: Gmail {e.resp.status}, retrying in {delay}s (attempt {attempt+1})")
                time.sleep(delay)
            else:
                raise
    return fn()


SUPPLIER_QUERY = {
    'Ashgrove':       'from:ashgrovecheese.com.au subject:{inv}',
    'Bega':           'from:noreplyBDD@bega.com.au subject:09{inv}',
    'Bidfood':        'from:bidfood.com.au subject:I{inv}',
    'IFP':            'from:mailforalex.com subject:{inv}',
    'Freshline':      'from:freshline.net.au subject:{inv}',
    'Juicy Isle':     'from:juicyisle.com.au subject:{inv}',
    'Goodman Fielder':'from:goodmanfielder.com.au subject:09{inv}',
    'Lactalis':       'from:parmalat.com.au subject:{inv}',
    'News Corp':      'from:circulation.news.com.au has:attachment',
    'Pandani Select': 'from:pandaniselect.com.au subject:{inv}',
    'Savour Foods':   'from:savourfoods.com.au subject:{inv}',
    'Scottsdale Pork':'from:fresho.com subject:Invoice subject:F{inv}',
    'Natures Foods':  'subject:"From Natures Foods" has:attachment',
    'Tas Bakeries':   'from:tasmanianbakeries.com.au subject:"Invoice from Tasmanian Bakeries" has:attachment',
    'Tas Gift Wrap':  'from:taswrap.com.au subject:INV{inv} has:attachment',
}

LABEL_QUERY = {
    'Ashgrove':       'label:Ashgrove-Cheese has:attachment',
    'Bega':           'label:Bega has:attachment',
    'Bidfood':        'label:Bidfood has:attachment',
    'IFP':            'label:Island-Fresh-Produce has:attachment',
    'Freshline':      'label:Freshline has:attachment',
    'Juicy Isle':     'label:Juicy-Isle has:attachment',
    'Goodman Fielder':'label:Goodman-Fielder has:attachment',
    'Lactalis':       'label:Lactalis has:attachment',
    'News Corp':      'label:News-Corp has:attachment',
    'Pandani Select': 'label:Pandani-Select has:attachment',
    'Savour Foods':   'label:Savour-Foods has:attachment',
    'Scottsdale Pork':'label:Fresho-Scottsdale has:attachment',
    'Tas Bakeries':   'label:Tasmanian-Bakeries-Natinal-Pies has:attachment',
    'Cartel & Co':    'label:Cartel-Co has:attachment',
    'PFD':            'label:PFD-Foods has:attachment',
    'Eden Foods':     'label:Eden-Foods has:attachment',
    'Olsen Eggs':     'label:Olsons-eggs has:attachment',
}

SUPPLIER_QUERY2 = {
    'Cripps Nu Bake':   'from:administration@cripps.com.au subject:"Cripps Invoice" has:attachment',
    'Wayside Butcher':  'from:wayside has:attachment',
    'Nichols Poultry':  'subject:"Nichols Poultry" has:attachment',
    'PFD':              'from:pfdfoods.com.au has:attachment',
    'Tasfresh':         'from:accounts.receivable@tasfresh.com.au has:attachment',
    'Scottsdale Pork':  'from:fresho.com subject:Invoice subject:F{inv} has:attachment',
    'Tas Bakeries':     'from:tasmanianbakeries.com.au has:attachment',
    'Sunrise Bakery':   'from:sunrise has:attachment',
    'Cartel & Co':      'from:cartelco.co has:attachment',
    'Horticultural L':  'from:dpritchard@hals.com.au subject:"HALS Invoice" has:attachment',
    'Natures Foods':    'subject:"From Natures Foods" has:attachment',
    'Olsen Eggs':       'subject:{inv} has:attachment',
    'Bega':             'from:noreplyBDD@bega.com.au has:attachment',
    'Bidfood':          'from:bidfood.com.au has:attachment',
    'Lactalis':         'from:parmalat.com.au has:attachment',
}


# Suppliers that only provide paper invoices (no email).
# These are skipped during Gmail search and marked PAPER INVOICE.
PAPER_SUPPLIERS = {
    'Wayside Butcher',
    'Sunrise Bakery',
    'Olsen Eggs',
    'Mountainvale',
    'Packings',
    'Licensed Socks',
    'Belle Esca',
    'Petuna Fisherie',
}

# Suppliers known to send weekly summary emails rather than per-invoice PDFs.
# For these, we require the invoice number to appear in the PDF text to avoid
# matching the summary (which has a fixed total unrelated to individual invoices).
SUMMARY_EMAIL_SUPPLIERS = {
    'Tas Bakeries',  # each delivery has its own invoice email; match by SO number in PDF
}

# Suppliers whose invoices arrive on EMAIL2 only.
# Skip email1 searches entirely for these to avoid wasting time.
EMAIL2_ONLY_SUPPLIERS = {
    'Tasfresh',
    'Nichols Poultry',
    'Horticultural L',
    'Cripps Nu Bake',
    'Natures Foods',
}

# Suppliers whose email invoice numbers have a prefix not in the TIR statement.
# e.g. TIR shows "51602430" but the email subject has "F51602430".
INV_EMAIL_PREFIX = {
    'Scottsdale Pork': 'F',
}

# Suppliers where the invoice number is in the attachment filename, not the subject.
# These need attachment-level matching instead of subject-based search.
ATTACHMENT_MATCH_SUPPLIERS = {
    'Nichols Poultry',
    'Natures Foods',
    'Tasfresh',
}

# Suppliers where the TIR invoice number differs from the email invoice number.
# Match by sender and approximate amount only — skip invoice-number-based fallback queries.
AMOUNT_MATCH_SUPPLIERS = {
    'Horticultural L',
    'PFD',  # TIR uses LT numbers, PFD emails use VT numbers — different ID systems, match by amount only
}

# Suppliers that send one weekly invoice covering multiple TIR line items.
# All TIR amounts for the same (supplier, date) are summed and compared against
# the single email invoice total. GST is distributed proportionally.
WEEKLY_INVOICE_SUPPLIERS = {
    'Cripps Nu Bake',
}

# Suppliers where the TIR invoice number has a prefix that should be stripped
# before searching within PDF text (e.g. TIR "NV118608" → "118608" matches "SO118608" in PDF).
INV_TIR_STRIP_PREFIX = {
    'Tas Bakeries',
}


def get_api_key():
    """Get Anthropic API key from env or Streamlit secrets."""
    key = os.environ.get("ANTHROPIC_API_KEY")
    if key:
        return key
    try:
        import streamlit as st
        return st.secrets["ANTHROPIC_API_KEY"]
    except Exception:
        pass
    raise ValueError("ANTHROPIC_API_KEY not found in environment or Streamlit secrets")


def _refresh_creds(creds):
    """Force-refresh OAuth credentials. Always refreshes if a refresh_token exists,
    regardless of the expiry field — the stored expiry may be stale or unparsed."""
    if creds.refresh_token:
        try:
            creds.refresh(Request())
            log.info("Gmail OAuth token refreshed successfully")
        except Exception as e:
            log.warning(f"Token refresh failed: {type(e).__name__}: {e}")
            # If refresh fails but token might still work, continue anyway
    return creds


def get_service(token_path):
    """Build Gmail API service from a token file path."""
    creds = Credentials.from_authorized_user_file(token_path)
    _refresh_creds(creds)
    return build('gmail', 'v1', credentials=creds)


def get_service_from_json(token_data):
    """Build Gmail API service from token JSON string or dict."""
    if isinstance(token_data, dict):
        info = dict(token_data)
    elif isinstance(token_data, str):
        # Strip wrapping quotes that Render/TOML may add
        s = token_data.strip().strip("'\"")
        info = json.loads(s)
    else:
        info = dict(token_data)
    creds = Credentials.from_authorized_user_info(info)
    _refresh_creds(creds)
    return build('gmail', 'v1', credentials=creds)


def get_services():
    """Get Gmail services from env vars, secrets, or local token files."""
    svc1, svc2 = None, None

    # Try env vars first (JSON strings)
    t1 = os.environ.get("TOKEN1")
    t2 = os.environ.get("TOKEN2")

    # Try Streamlit secrets
    if not t1:
        try:
            import streamlit as st
            t1 = st.secrets.get("TOKEN1")
            t2 = st.secrets.get("TOKEN2")
        except Exception:
            pass

    if t1:
        svc1 = get_service_from_json(t1)
        if t2:
            svc2 = get_service_from_json(t2)
        return svc1, svc2

    # Fall back to local token files
    token1_path = os.path.expanduser('~/Desktop/mts-gst-app/token.json')
    token2_path = os.path.expanduser('~/Desktop/mts-gst-app/token2.json')
    if os.path.exists(token1_path):
        svc1 = get_service(token1_path)
    if os.path.exists(token2_path):
        svc2 = get_service(token2_path)
    return svc1, svc2


def _gmail_search(svc, query, max_results=GMAIL_MAX_RESULTS):
    """Thread-safe, cached Gmail search. Avoids redundant queries for the same supplier."""
    key = (id(svc), query)
    with _gmail_lock:
        if key in _gmail_search_cache:
            return _gmail_search_cache[key]
        try:
            msgs = _retry(
                lambda: svc.users().messages().list(userId='me', q=query, maxResults=max_results).execute(),
                f"Gmail search '{query[:50]}'"
            ).get('messages', [])
            # Only cache successful results — don't cache errors as empty
            _gmail_search_cache[key] = msgs
            return msgs
        except Exception as e:
            log.warning(f"Gmail search failed: {type(e).__name__}: {e}")
            # Do NOT cache — transient errors should be retried on next call
            return []


def _gmail_get_message(svc, msg_id):
    """Thread-safe Gmail message fetch."""
    with _gmail_lock:
        return _retry(
            lambda: svc.users().messages().get(userId='me', id=msg_id, format='full').execute(),
            f"Gmail get message {msg_id}")


def _gmail_get_attachment(svc, msg_id, att_id):
    """Thread-safe Gmail attachment fetch."""
    with _gmail_lock:
        return _retry(
            lambda aid=att_id, mid=msg_id: svc.users().messages().attachments().get(
                userId='me', messageId=mid, id=aid).execute(),
            f"Gmail get attachment {msg_id}")


def read_pdfs(svc, msg_id):
    """Extract text from PDF attachments in a Gmail message. Results are cached by msg_id."""
    if msg_id in _pdf_text_cache:
        return _pdf_text_cache[msg_id]

    texts = []
    try:
        msg = _gmail_get_message(svc, msg_id)

        def scan(parts):
            for p in parts:
                if p.get('parts'):
                    scan(p['parts'])
                if p.get('filename', '').lower().endswith('.pdf'):
                    aid = p.get('body', {}).get('attachmentId')
                    if aid:
                        att = _gmail_get_attachment(svc, msg_id, aid)
                        data = base64.urlsafe_b64decode(att['data'])
                        with pdfplumber.open(io.BytesIO(data)) as pdf:
                            t = '\n'.join(pg.extract_text() or '' for pg in pdf.pages)
                            if t.strip():
                                texts.append(t)
        payload = msg.get('payload', {})
        scan(payload.get('parts', []))

        # Handle single-part emails where the PDF IS the top-level payload
        # (e.g. Lactalis sends emails where the entire payload is the PDF attachment)
        if not texts and payload.get('filename', '').lower().endswith('.pdf'):
            aid = payload.get('body', {}).get('attachmentId')
            if aid:
                att = _gmail_get_attachment(svc, msg_id, aid)
                data = base64.urlsafe_b64decode(att['data'])
                with pdfplumber.open(io.BytesIO(data)) as pdf:
                    t = '\n'.join(pg.extract_text() or '' for pg in pdf.pages)
                    if t.strip():
                        texts.append(t)

    except (HttpError, Exception) as e:
        log.warning(f"read_pdfs({msg_id}): {type(e).__name__}: {e}")

    _pdf_text_cache[msg_id] = texts
    return texts


def _extract_with_regex(text, supplier=None):
    """Extract total and GST using regex for known invoice formats.

    Returns (total, gst) or (None, None) if no pattern matches.
    Much faster and more reliable than LLM extraction for structured invoices.
    """
    if supplier == 'Cripps Nu Bake':
        # Last page has "INVOICE TOTAL: 2903.93" or "INVOICE TOTAL: 227.42-" (credit)
        m = re.search(r'INVOICE TOTAL:\s*([\d,]+\.\d{2})(-)?', text)
        if m:
            total = float(m.group(1).replace(',', ''))
            if m.group(2):
                total = -total
            # GST from the last "Total ..." data row — second-to-last decimal number
            gst = 0.0
            for line_m in re.finditer(r'^Total\s+(.+)$', text, re.MULTILINE):
                nums = re.findall(r'-?[\d,]+\.\d{2}', line_m.group(1))
                if len(nums) >= 2:
                    gst = float(nums[-2].replace(',', ''))
            return total, gst

    if supplier == 'Tas Bakeries':
        # "Total $412.59" and "GST Total $37.53"
        total_m = re.search(r'^\s*Total\s+\$([\d,]+\.\d{2})\s*$', text, re.MULTILINE)
        gst_m = re.search(r'GST Total\s+\$([\d,]+\.\d{2})', text)
        if total_m:
            total = float(total_m.group(1).replace(',', ''))
            gst = float(gst_m.group(1).replace(',', '')) if gst_m else 0.0
            return total, gst

    if supplier == 'PFD':
        # "ORDER TOTAL (GST Included) $634.10" and "TOTAL GST $57.34"
        total_m = re.search(r'ORDER TOTAL \(GST Included\)\s+\$?([\d,]+\.\d{2})', text)
        gst_m = re.search(r'TOTAL GST\s+\$?([\d,]+\.\d{2})', text)
        if total_m:
            total = float(total_m.group(1).replace(',', ''))
            gst = float(gst_m.group(1).replace(',', '')) if gst_m else 0.0
            return total, gst

    if supplier == 'Lactalis':
        # "TOTAL AMOUNT: $ 119.74" and "Total of taxable supplies 83.20"
        total_m = re.search(r'TOTAL AMOUNT:\s*\$\s*([\d,]+\.\d{2})', text)
        if total_m:
            total = float(total_m.group(1).replace(',', ''))
            # "Total of taxable supplies" is GST-inclusive; GST = taxable / 11
            taxable_m = re.search(r'Total of taxable supplies\s+([\d,]+\.\d{2})', text)
            if taxable_m:
                taxable = float(taxable_m.group(1).replace(',', ''))
                gst = round(taxable / 11, 2)
            else:
                gst = 0.0
            return total, gst

    if supplier == 'Horticultural L':
        # "Total 165.84 16.58 182.42" columns are Ex GST | GST | Inc GST
        m = re.search(r'^Total\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})', text, re.MULTILINE)
        if m:
            gst = float(m.group(2).replace(',', ''))
            total = float(m.group(3).replace(',', ''))
            return total, gst

    return None, None


def extract_gst_and_total(text, client, supplier=None):
    """Extract total and GST from invoice text.

    Tries deterministic regex patterns first (fast, reliable for known formats),
    then falls back to Claude Haiku for unknown formats.
    """
    # Try regex-based extraction first
    total, gst = _extract_with_regex(text, supplier)
    if total is not None:
        log.info(f"extract_gst_and_total: regex matched for {supplier}: total={total}, gst={gst}")
        return total, gst

    # Fall back to Claude Haiku
    def _call():
        return client.messages.create(
            model='claude-haiku-4-5-20251001', max_tokens=150,
            messages=[{'role': 'user', 'content':
                f'From this invoice extract TOTAL AMOUNT DUE and TOTAL GST INCLUDED. '
                f'Return ONLY JSON: {{"total":0.00,"gst":0.00}}\n\n{text[-2000:]}'}])

    r = _retry(_call, "Claude extract GST")
    t = re.sub(r'```json|```', '', r.content[0].text).strip()
    d = json.loads(t[t.find('{'):t.rfind('}') + 1])
    total = d.get('total')
    gst = d.get('gst')
    # Default GST to 0.0 when we successfully extracted a total — makes it clear
    # the invoice was found but has no/zero GST, vs invoice not found (None).
    if total is not None and gst is None:
        gst = 0.0
    return total, gst


def search_and_verify(svc, query, tir_amount, inv_no, client, require_inv_in_text=False, supplier=None):
    """Search Gmail for an invoice and verify the amount matches TIR.

    If require_inv_in_text is True, the invoice number must appear in the PDF
    text for a match to be accepted. This prevents summary emails (with a fixed
    total) from being mistaken for individual invoices.
    For suppliers in INV_TIR_STRIP_PREFIX, the numeric part of inv_no is also tried.
    """
    msgs = _gmail_search(svc, query)

    # Track first mismatch as fallback — don't stop on first non-matching PDF,
    # keep trying other messages to find the right invoice.
    first_mismatch = None
    for m in msgs:
        for txt in read_pdfs(svc, m['id']):
            if require_inv_in_text and inv_no:
                # Also try the numeric-only part of the invoice number
                # (e.g. TIR "NV118608" → "118608" matches "SO118608" in Tas Bakeries PDFs)
                inv_num = re.sub(r'^[A-Za-z]+', '', inv_no)
                if inv_no not in txt and inv_num not in txt:
                    log.info(f"search_and_verify: skipping msg {m['id']} — inv {inv_no}/{inv_num} not found in PDF text")
                    continue
            try:
                t, g = extract_gst_and_total(txt, client, supplier=supplier)
                if t is not None and abs(abs(t) - abs(tir_amount)) < 1.00:
                    return g, t, 'VERIFIED ✓'
                elif t is not None and first_mismatch is None:
                    first_mismatch = (g, t, f'AMOUNT MISMATCH (PDF:{t} TIR:{tir_amount})')
            except (json.JSONDecodeError, KeyError, IndexError) as e:
                log.warning(f"search_and_verify: parse error for msg {m['id']}: {e}")
            except Exception as e:
                log.warning(f"search_and_verify: unexpected error for msg {m['id']}: {type(e).__name__}: {e}")
    if first_mismatch:
        return first_mismatch
    return None, None, None


def _filename_matches_inv(fname, inv_no):
    """Check if a filename contains the invoice number.

    Handles patterns like:
      InvoiceFRS2002177B.pdf → extracts 2002177, matches inv_no '2002177'
      InvoiceLAU4526767.pdf  → extracts 4526767, matches inv_no '4526767'
      Invoice554148.P        → extracts 554148,  matches inv_no '554148'
      INV00024807.p          → zero-stripped match, matches inv_no '24807'
    """
    # Extract all digit sequences from filename
    digits = re.findall(r'\d+', fname)
    # Exact match against any extracted number
    if inv_no in digits:
        return True
    # Zero-stripped comparison (handles zero-padded numbers like 00024807 matching 24807)
    inv_stripped = inv_no.lstrip('0') or inv_no
    if any(d.lstrip('0') == inv_stripped for d in digits):
        return True
    return False


def search_and_verify_by_attachment(svc, query, tir_amount, inv_no, client, supplier=None):
    """Search Gmail and match invoice by attachment filename (e.g. Invoice554148.P).

    Used for suppliers like Nichols Poultry where every email has the same subject
    and the invoice number only appears in the attachment filename.
    """
    msgs = _gmail_search(svc, query)
    log.info(f"search_and_verify_by_attachment: query='{query[:60]}' inv={inv_no} found {len(msgs)} messages")

    first_mismatch = None
    for m in msgs:
        msg_id = m['id']
        try:
            msg = _gmail_get_message(svc, msg_id)
        except Exception as e:
            log.warning(f"search_and_verify_by_attachment({msg_id}): {type(e).__name__}: {e}")
            continue

        # Scan all parts (including top-level payload) for attachments containing invoice number
        matches = []
        all_filenames = []  # for debug logging

        def scan(parts):
            for p in parts:
                if p.get('parts'):
                    scan(p['parts'])
                fname = p.get('filename', '')
                if fname:
                    all_filenames.append(fname)
                    if _filename_matches_inv(fname, inv_no):
                        aid = p.get('body', {}).get('attachmentId')
                        if aid:
                            matches.append((fname, aid))

        # Scan nested parts AND the top-level payload itself
        payload = msg.get('payload', {})
        if payload.get('parts'):
            scan(payload['parts'])
        # Some emails have the attachment at the top level (single-part)
        if payload.get('filename') and _filename_matches_inv(payload.get('filename', ''), inv_no):
            aid = payload.get('body', {}).get('attachmentId')
            if aid:
                matches.append((payload['filename'], aid))

        if not matches and all_filenames:
            log.info(f"  msg {msg_id}: no inv_no match in filenames: {all_filenames[:5]}")

        for fname, aid in matches:
            try:
                att = _gmail_get_attachment(svc, msg_id, aid)
                data = base64.urlsafe_b64decode(att['data'])
                with pdfplumber.open(io.BytesIO(data)) as pdf:
                    txt = '\n'.join(pg.extract_text() or '' for pg in pdf.pages)
                if not txt.strip():
                    continue
                t, g = extract_gst_and_total(txt, client, supplier=supplier)
                if t is not None and abs(abs(t) - abs(tir_amount)) < 1.00:
                    return g, t, 'VERIFIED ✓'
                elif t is not None and first_mismatch is None:
                    first_mismatch = (g, t, f'AMOUNT MISMATCH (PDF:{t} TIR:{tir_amount})')
            except Exception as e:
                log.warning(f"search_and_verify_by_attachment: error processing {fname} in {msg_id}: {e}")

    if first_mismatch:
        return first_mismatch
    return None, None, None


def search_and_verify_weekly(svc, query, weekly_sum, client, supplier=None):
    """Search Gmail for a weekly invoice and verify the total matches TIR weekly sum.

    Unlike search_and_verify, this sums ALL PDF totals within each email message
    before comparing. This handles suppliers like Cripps that send a main invoice
    plus a credit note as separate attachments in the same email.
    """
    msgs = _gmail_search(svc, query)
    first_mismatch = None

    for m in msgs:
        texts = read_pdfs(svc, m['id'])
        if not texts:
            continue

        msg_total = 0.0
        msg_gst = 0.0
        all_ok = True

        for txt in texts:
            try:
                t, g = extract_gst_and_total(txt, client, supplier=supplier)
                if t is not None:
                    msg_total += t
                    msg_gst += (g or 0.0)
                else:
                    all_ok = False
            except Exception as e:
                log.warning(f"search_and_verify_weekly: extract error for msg {m['id']}: {e}")
                all_ok = False

        if not all_ok or msg_total == 0.0 and len(texts) > 0:
            # If we couldn't extract from all PDFs, also try each PDF individually
            # (single-PDF emails still work)
            for txt in texts:
                try:
                    t, g = extract_gst_and_total(txt, client, supplier=supplier)
                    if t is not None and abs(abs(t) - abs(weekly_sum)) < 1.00:
                        return g, t, 'VERIFIED ✓'
                except Exception:
                    pass
            continue

        if abs(abs(msg_total) - abs(weekly_sum)) < 1.00:
            return msg_gst, msg_total, 'VERIFIED ✓'

        # If sum doesn't match but there are multiple PDFs (e.g. invoice + credit note),
        # try matching individual PDFs. The credit note may be a separate TIR line item
        # on a different date, so the TIR weekly sum may only cover the main invoice.
        if len(texts) > 1:
            for txt in texts:
                try:
                    t, g = extract_gst_and_total(txt, client, supplier=supplier)
                    if t is not None and t > 0 and abs(abs(t) - abs(weekly_sum)) < 1.00:
                        return g, t, 'VERIFIED ✓'
                except Exception:
                    pass

        if first_mismatch is None:
            first_mismatch = (msg_gst, msg_total,
                              f'AMOUNT MISMATCH (PDF:{msg_total:.2f} TIR:{weekly_sum})')

    if first_mismatch:
        return first_mismatch
    return None, None, None


def parse_tir_pdf(pdf_bytes, client):
    """Parse a TIR statement PDF into a list of (date, supplier, inv_no, amount) tuples."""
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        full_text = '\n'.join(pg.extract_text() or '' for pg in pdf.pages)

    def _call():
        return client.messages.create(
            model='claude-sonnet-4-20250514', max_tokens=8000,
            messages=[{'role': 'user', 'content':
                'Extract ALL invoice line items from this TIR statement. '
                'For each line return: date (DD/MM/YYYY), supplier name, invoice number, and amount (as a number, negative for credits). '
                'Return ONLY a JSON array of arrays: [["06/03/2026","Supplier Name","INV123",427.62], ...]\n'
                'Include every single line item. Do not skip any.\n\n'
                f'{full_text}'}])

    r = _retry(_call, "Claude parse TIR PDF")
    text = re.sub(r'```json|```', '', r.content[0].text).strip()
    data = json.loads(text[text.find('['):text.rfind(']') + 1])
    return [(row[0], row[1], str(row[2]), float(row[3])) for row in data]


def _process_one_invoice(date, supplier, inv_no, tir_amount, svc1, svc2, client):
    """Process a single invoice — search Gmail, extract GST, verify amount.

    Returns (date, supplier, inv_no, tir_amount, gst, inv_total, status).
    """
    gst = None
    inv_total = None
    status = 'NOT FOUND'

    needs_inv_check = supplier in SUMMARY_EMAIL_SUPPLIERS
    email_inv_no = INV_EMAIL_PREFIX.get(supplier, '') + inv_no

    # Try email 1
    if supplier not in EMAIL2_ONLY_SUPPLIERS:
        q_specific = SUPPLIER_QUERY.get(supplier, '').replace('{inv}', inv_no)

        if supplier in ATTACHMENT_MATCH_SUPPLIERS and q_specific:
            g, t, s = search_and_verify_by_attachment(svc1, q_specific, tir_amount, inv_no, client, supplier=supplier)
            if s and 'VERIFIED' in s:
                gst, inv_total, status = g, t, s
            elif s:
                gst, inv_total, status = g, t, s
        else:
            queries_email1 = [q for q in [q_specific, f'subject:{email_inv_no} has:attachment -from:tir.com.au'] if q.strip()]

            lq = LABEL_QUERY.get(supplier)
            if lq:
                g, t, s = search_and_verify(svc1, lq, tir_amount, inv_no, client,
                                             require_inv_in_text=needs_inv_check, supplier=supplier)
                if s and 'VERIFIED' in s:
                    gst, inv_total, status = g, t, 'VERIFIED ✓ (label)'
                elif s and gst is None:
                    gst, inv_total, status = g, t, s

            for q in queries_email1:
                if 'VERIFIED' in status:
                    break
                g, t, s = search_and_verify(svc1, q, tir_amount, inv_no, client,
                                             require_inv_in_text=needs_inv_check, supplier=supplier)
                if s and 'VERIFIED' in s:
                    gst, inv_total, status = g, t, s
                elif s:
                    gst, inv_total, status = g, t, s

    # Try email 2 if not verified yet
    if 'VERIFIED' not in status and svc2:
        q2_tmpl = SUPPLIER_QUERY2.get(supplier, f'subject:{email_inv_no} has:attachment')
        q2_specific = q2_tmpl.replace('{inv}', inv_no)

        if supplier in ATTACHMENT_MATCH_SUPPLIERS:
            g, t, s = search_and_verify_by_attachment(svc2, q2_specific, tir_amount, inv_no, client, supplier=supplier)
            if s and 'VERIFIED' in s:
                gst, inv_total, status = g, t, 'VERIFIED ✓ (email2)'
            elif s and 'MISMATCH' not in status:
                gst, inv_total, status = g, t, f'MISMATCH email2 (PDF:{t} TIR:{tir_amount})'
        elif supplier in AMOUNT_MATCH_SUPPLIERS:
            g, t, s = search_and_verify(svc2, q2_specific, tir_amount, inv_no, client, supplier=supplier)
            if s and 'VERIFIED' in s:
                gst, inv_total, status = g, t, 'VERIFIED ✓ (email2)'
            elif s and 'MISMATCH' not in status:
                gst, inv_total, status = g, t, f'MISMATCH email2 (PDF:{t} TIR:{tir_amount})'
        else:
            queries_email2 = [q2_specific, f'subject:{email_inv_no} has:attachment -from:tir.com.au']
            for q in queries_email2:
                g, t, s = search_and_verify(svc2, q, tir_amount, inv_no, client,
                                             require_inv_in_text=needs_inv_check, supplier=supplier)
                if s and 'VERIFIED' in s:
                    gst, inv_total, status = g, t, 'VERIFIED ✓ (email2)'
                    break
                elif s and 'MISMATCH' not in status:
                    gst, inv_total, status = g, t, f'MISMATCH email2 (PDF:{t} TIR:{tir_amount})'

    return (date, supplier, inv_no, tir_amount, gst, inv_total, status)


def reconcile(tir_data, svc1, svc2, api_key, progress_callback=None):
    """Run reconciliation on TIR data. Returns list of result tuples.

    Processes invoices in parallel using a thread pool for speed.
    progress_callback(i, total, supplier, inv_no, status) is called after each invoice.
    """
    client = anthropic.Anthropic(api_key=api_key)
    # Clear caches from any previous run
    _pdf_text_cache.clear()
    _gmail_search_cache.clear()

    total_count = len(tir_data)

    # --- Pre-process weekly invoice suppliers ---
    _weekly_groups = {}
    for date, supplier, inv_no, tir_amount in tir_data:
        if supplier in WEEKLY_INVOICE_SUPPLIERS:
            _weekly_groups.setdefault((supplier, date), []).append((inv_no, tir_amount))

    _weekly_cache = {}
    for (supplier, date), items in _weekly_groups.items():
        weekly_sum = sum(amt for _, amt in items)
        svc = svc2 if supplier in EMAIL2_ONLY_SUPPLIERS and svc2 else svc1
        q = SUPPLIER_QUERY2.get(supplier, SUPPLIER_QUERY.get(supplier, ''))
        if svc and q:
            g, t, s = search_and_verify_weekly(svc, q, weekly_sum, client, supplier=supplier)
            if s:
                tag = 'email2' if svc is svc2 else 'email1'
                status_str = f'VERIFIED ✓ ({tag}, weekly)' if 'VERIFIED' in s else s
                _weekly_cache[(supplier, date)] = (g, t, weekly_sum, status_str)

    # --- Build results: handle paper + weekly synchronously, batch-parallelize the rest ---
    results_map = {}  # index -> list of result tuples
    _progress_lock = threading.Lock()
    _progress_counter = [0]

    def _report_progress(supplier, inv_no, status):
        with _progress_lock:
            _progress_counter[0] += 1
            if progress_callback:
                progress_callback(_progress_counter[0], total_count, supplier, inv_no, status)

    # Separate items into instant (no API) and parallel (needs API) buckets
    parallel_items = []  # (index, date, supplier, inv_no, tir_amount)

    for i, (date, supplier, inv_no, tir_amount) in enumerate(tir_data):

        # TIR internal charges — not supplier invoices (e.g. TIR-WklyPoster, TIR-MthCat)
        if supplier.startswith('TIR-') or supplier.startswith('TIR '):
            results_map[i] = [(date, supplier, inv_no, tir_amount, None, None, 'TIR INTERNAL')]
            _report_progress(supplier, inv_no, 'TIR INTERNAL')
            continue

        # Paper suppliers — no API calls needed
        if supplier in PAPER_SUPPLIERS:
            results_map[i] = [(date, supplier, inv_no, tir_amount, None, None, 'PAPER INVOICE')]
            _report_progress(supplier, inv_no, 'PAPER INVOICE')
            continue

        # Weekly suppliers — already resolved in pre-processing
        if supplier in WEEKLY_INVOICE_SUPPLIERS:
            rows = []
            cached = _weekly_cache.get((supplier, date))
            status = cached[3] if cached else 'NOT FOUND'
            rows.append((date, supplier, inv_no, tir_amount, None, None, status))

            group_items = _weekly_groups.get((supplier, date), [])
            last_inv = group_items[-1][0] if group_items else None
            if inv_no == last_inv and cached:
                weekly_gst, inv_total, weekly_sum, wk_status = cached
                rows.append((date, supplier, 'WEEKLY TOTAL', weekly_sum, weekly_gst, inv_total, wk_status))

            results_map[i] = rows
            _report_progress(supplier, inv_no, status)
            continue

        parallel_items.append((i, date, supplier, inv_no, tir_amount))

    # Process remaining invoices in batches to limit peak memory on Render free tier
    for batch_start in range(0, len(parallel_items), BATCH_SIZE):
        batch = parallel_items[batch_start:batch_start + BATCH_SIZE]
        futures = {}

        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            for idx, date, supplier, inv_no, tir_amount in batch:
                future = executor.submit(
                    _process_one_invoice, date, supplier, inv_no, tir_amount, svc1, svc2, client)
                futures[future] = (idx, date, supplier, inv_no, tir_amount)

            for future in concurrent.futures.as_completed(futures):
                idx, date, supplier, inv_no, tir_amount = futures[future]
                try:
                    result = future.result()
                    results_map[idx] = [result]
                    _report_progress(supplier, inv_no, result[6])
                except Exception as e:
                    # Error recovery: skip failed invoice instead of crashing
                    log.error(f"Error processing {supplier} {inv_no}: {type(e).__name__}: {e}")
                    results_map[idx] = [(date, supplier, inv_no, tir_amount, None, None, f'ERROR: {type(e).__name__}')]
                    _report_progress(supplier, inv_no, 'ERROR')

        # Force garbage collection between batches to free memory
        gc.collect()

    # Reassemble results in original TIR order
    results = []
    for i in range(len(tir_data)):
        results.extend(results_map.get(i, []))

    return results


def build_excel(results, title_date=''):
    """Build Excel workbook from results. Returns bytes."""
    total_count = len(results)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'GST Reconciliation'
    ws.merge_cells('A1:G1')
    ws['A1'] = f'IGA Campbell Town — TIR GST Reconciliation {title_date}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:G2')
    ws['A2'] = 'GST extracted from supplier invoice PDFs — verified by matching TIR total amount'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.append([])
    ws.append(['Date', 'Supplier', 'Invoice No.', 'TIR Amount ($)', 'Invoice Total ($)', 'Actual GST ($)', 'Status'])
    for c in ws[4]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='1a2d45')
        c.alignment = Alignment(horizontal='center')

    green = PatternFill('solid', start_color='E8F5E9')
    orange = PatternFill('solid', start_color='FFF3E0')
    red = PatternFill('solid', start_color='FFEBEE')
    grey = PatternFill('solid', start_color='F5F5F5')
    purple = PatternFill('solid', start_color='E8D5F5')
    total_gst = 0
    verified_count = 0

    # Track which (supplier, date) weekly groups have been counted as verified
    _weekly_verified = set()

    for date, supplier, inv_no, tir_amt, gst, inv_total, status in results:
        is_weekly_total = inv_no == 'WEEKLY TOTAL'
        # Show 0.00 for GST when invoice was found (gst is not None), blank when not found
        gst_display = gst if gst is not None else ''
        inv_display = inv_total if inv_total is not None else ''
        ws.append([date, supplier, inv_no, tir_amt, inv_display, gst_display, status])
        row = ws[ws.max_row]
        if is_weekly_total:
            # Bold the weekly total row; don't count towards total_count
            for c in row:
                c.fill = green if 'VERIFIED' in status else orange
                c.font = Font(bold=True)
            if gst is not None:
                total_gst += gst
            if 'VERIFIED' in status:
                _weekly_verified.add((supplier, date))
        elif 'VERIFIED' in status:
            for c in row:
                c.fill = green
            if gst is not None:
                total_gst += gst
            if (supplier, date) not in _weekly_verified:
                # Normal verified row (not part of a weekly group)
                verified_count += 1
        elif 'TIR INTERNAL' in status:
            for c in row:
                c.fill = purple
        elif 'PAPER INVOICE' in status:
            for c in row:
                c.fill = grey
        elif 'MISMATCH' in status:
            for c in row:
                c.fill = orange
        else:
            for c in row:
                c.fill = red

    # Count each verified weekly group as one
    verified_count += len(_weekly_verified)
    # Exclude WEEKLY TOTAL synthetic rows from TIR total sum
    tir_total = sum(r[3] for r in results if r[2] != 'WEEKLY TOTAL')
    # Exclude WEEKLY TOTAL rows from the total invoice count
    invoice_count = sum(1 for r in results if r[2] != 'WEEKLY TOTAL')

    ws.append([])
    ws.append(['', '', 'TOTALS', tir_total, '', round(total_gst, 2),
               f'{verified_count}/{invoice_count} verified'])
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for col, w in zip('ABCDEFG', [14, 20, 16, 16, 16, 16, 28]):
        ws.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), verified_count, invoice_count, total_gst


# CLI entry point
if __name__ == '__main__':
    api_key = get_api_key()
    svc1, svc2 = get_services()

    # Hardcoded TIR for CLI testing — in production, use parse_tir_pdf()
    TIR = [
        ('06/03/2026','Ashgrove','723418',427.62),('06/03/2026','Ashgrove','724095',147.44),
        ('06/03/2026','Ashgrove','724704',558.80),('06/03/2026','Ashgrove','724705',293.13),
        ('06/03/2026','Bega','25904255',728.03),('06/03/2026','Bega','25911322',1165.66),
        ('06/03/2026','Bega','25918561',990.95),('06/03/2026','Bega','25934042',843.02),
        ('06/03/2026','Bega','77007488',-620.05),('06/03/2026','Bidfood','69543372',220.27),
        ('06/03/2026','Bidfood','C6974657',42.12),('06/03/2026','Cartel & Co','IOA86090',530.13),
        ('06/03/2026','Cripps Nu Bake','925206',2614.83),('06/03/2026','Cripps Nu Bake','926026',2676.54),
        ('06/03/2026','Eden Foods','RIN74955',2883.09),('06/03/2026','Eden Foods','RIN75028',74.79),
        ('06/03/2026','Eden Foods','RIN75153',115.74),('06/03/2026','Freshline','S338646',1641.11),
        ('06/03/2026','Freshline','S339355',2194.47),('06/03/2026','Freshline','S340009',2168.97),
        ('06/03/2026','Goodman Fielder','98633224',526.31),('06/03/2026','Horticultural L','241393-1',513.89),
        ('06/03/2026','IFP','1515901',-160.17),('06/03/2026','IFP','1516070',-64.57),
        ('06/03/2026','IFP','1516240',60.00),('06/03/2026','IFP','1516368',120.00),
        ('06/03/2026','IFP','1516451',-13.50),('06/03/2026','IFP','1516461',-22.50),
        ('06/03/2026','IFP','1516478',22.50),('06/03/2026','IFP','1516503',-40.56),
        ('06/03/2026','IFP','M261862',433.26),('06/03/2026','IFP','M262113',803.03),
        ('06/03/2026','IFP','M262249',524.47),('06/03/2026','IFP','M262476',2174.10),
        ('06/03/2026','IFP','M262656',1629.22),('06/03/2026','IFP','M262673',506.62),
        ('06/03/2026','IFP','M262703',176.70),('06/03/2026','IFP','M262812',1575.17),
        ('06/03/2026','IFP','M262834',167.05),('06/03/2026','IFP','M262839',544.47),
        ('06/03/2026','IFP','M262855',41.90),('06/03/2026','IFP','M263072',1630.77),
        ('06/03/2026','Juicy Isle','1444158',3314.60),('06/03/2026','Lactalis','42964661',181.18),
        ('06/03/2026','Lactalis','43063841',441.78),('06/03/2026','Licensed Socks','I0153649',653.40),
        ('06/03/2026','News Corp','8048633',549.08),('06/03/2026','Nichols Poultry','553981',487.27),
        ('06/03/2026','Nichols Poultry','554148',515.44),('06/03/2026','Pandani Select','378328',443.87),
        ('06/03/2026','Pandani Select','378477',249.26),('06/03/2026','Petuna Fisherie','20078404',143.88),
        ('06/03/2026','PFD','LT658333',506.05),('06/03/2026','PFD','LT670256',-39.71),
        ('06/03/2026','Savour Foods','SAV55775',493.26),('06/03/2026','Scottsdale Pork','50770801',694.46),
        ('06/03/2026','Scottsdale Pork','50861583',699.03),('06/03/2026','Sunrise Bakery','912070',222.05),
        ('06/03/2026','Tas Bakeries','NV118186',786.66),('06/03/2026','Tas Bakeries','NV118503',205.64),
        ('06/03/2026','Tasfresh','118664',-269.06),('06/03/2026','Tasfresh','2001709',279.10),
        ('06/03/2026','Tasfresh','2001714',50.00),('06/03/2026','Tasfresh','2001946',255.90),
        ('06/03/2026','Tasfresh','4526767',1061.28),('06/03/2026','Tasfresh','4526845',403.51),
        ('06/03/2026','Tasfresh','4527785',309.84),('06/03/2026','Tasfresh','4527940',1023.63),
        ('06/03/2026','Wayside Butcher','5735',2177.56),('06/03/2026','Wayside Butcher','5748',2389.72),
        ('13/03/2026','Ashgrove','725329',201.98),('13/03/2026','Ashgrove','726027',347.17),
        ('13/03/2026','Ashgrove','726936',653.63),('13/03/2026','Ashgrove','726937',404.27),
        ('13/03/2026','Bega','25949221',1014.21),('13/03/2026','Bega','25956542',918.53),
        ('13/03/2026','Bega','25956543',1714.83),('13/03/2026','Bega','25977037',475.17),
        ('13/03/2026','Bega','77036666',-935.44),('13/03/2026','Bidfood','69618480',726.82),
        ('13/03/2026','Cripps Nu Bake','926886',2903.39),('13/03/2026','Eden Foods','RIN75406',2780.54),
        ('13/03/2026','Freshline','S340947',1504.14),('13/03/2026','Goodman Fielder','98650450',539.82),
        ('13/03/2026','IFP','M263173',162.10),('13/03/2026','IFP','M263187',1067.99),
        ('13/03/2026','IFP','M263348',1149.39),('13/03/2026','IFP','M263496',1396.67),
        ('13/03/2026','IFP','M263663',2035.42),('13/03/2026','IFP','M263821',1261.29),
        ('13/03/2026','IFP','M264035',1494.77),('13/03/2026','Juicy Isle','1445371',7683.71),
        ('13/03/2026','Juicy Isle','1445498',614.89),('13/03/2026','Juicy Isle','R1445371',-745.20),
        ('13/03/2026','Momentum Foods','10071798',643.88),('13/03/2026','Natures Foods','24807',476.80),
        ('13/03/2026','News Corp','8056663',477.07),('13/03/2026','Nichols Poultry','554360',260.42),
        ('13/03/2026','Nichols Poultry','554490',166.14),('13/03/2026','Olsen Eggs','62514',414.00),
        ('13/03/2026','Olsen Eggs','62628',207.00),('13/03/2026','Pandani Select','378602',255.63),
        ('13/03/2026','Pandani Select','378760',291.68),('13/03/2026','Pandani Select','378854',346.88),
        ('13/03/2026','Petuna Fisherie','20078566',203.00),('13/03/2026','PFD','LT731916',1035.00),
        ('13/03/2026','Scottsdale Pork','50967307',653.86),('13/03/2026','Scottsdale Pork','51128792',721.04),
        ('13/03/2026','Sunrise Bakery','912209',177.55),('13/03/2026','Tas Bakeries','NV118736',426.47),
        ('13/03/2026','Tas Bakeries','NV119075',538.28),('13/03/2026','Tasfresh','2002177',259.35),
        ('13/03/2026','Tasfresh','2002387',270.25),('13/03/2026','Tasfresh','4529196',1032.76),
        ('13/03/2026','Tasfresh','4529197',286.88),('13/03/2026','Tasfresh','4529885',541.49),
        ('13/03/2026','Tasfresh','4529936',1422.92),('13/03/2026','Tasfresh','9185051',444.48),
        ('13/03/2026','Wayside Butcher','5764',1861.47),('13/03/2026','Wayside Butcher','5765',818.90),
        ('13/03/2026','Wayside Butcher','5781',2115.93),('13/03/2026','Wayside Butcher','5787',108.94),
    ]

    def cli_progress(i, total, supplier, inv_no, status):
        print(f'[{i}/{total}] {supplier} {inv_no} → {status}')

    results = reconcile(TIR, svc1, svc2, api_key, progress_callback=cli_progress)
    excel_bytes, verified, total, total_gst = build_excel(results, 'Week Ended 13/03/2026')

    out = os.path.expanduser('~/Desktop/MTS_GST_13Mar2026.xlsx')
    with open(out, 'wb') as f:
        f.write(excel_bytes)
    print(f'\n✅ DONE! Saved: {out}')
    print(f'Verified: {verified}/{total} | Total GST: ${total_gst:.2f}')
